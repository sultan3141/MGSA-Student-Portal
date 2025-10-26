from rest_framework import permissions, status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from accounts.models import User
from posts.models import Post, Like
from resources.models import Resource
from tutorials.models import Tutorial, TutorialRegistration
from .models import Feedback, UserActivity, SystemAnalytics
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import io

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def dashboard_stats(request):
    if request.user.role not in ['Executive', 'Admin']:
        return Response({
            'success': False,
            'message': 'Access denied. Executive or Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # User statistics
    user_stats = User.objects.filter(is_active=True).aggregate(
        total_users=Count('id'),
        students=Count('id', filter=Q(role='Student')),
        executives=Count('id', filter=Q(role='Executive')),
        admins=Count('id', filter=Q(role='Admin'))
    )
    
    # Users by department
    users_by_department = User.objects.filter(
        is_active=True, role='Student'
    ).values('department').annotate(count=Count('id')).order_by('-count')
    
    # Users by year
    users_by_year = User.objects.filter(
        is_active=True, role='Student'
    ).values('year_of_study').annotate(count=Count('id')).order_by('year_of_study')
    
    # Users by zone
    users_by_zone = User.objects.filter(
        is_active=True, role='Student'
    ).values('zone').annotate(count=Count('id')).order_by('-count')
    
    # Post statistics
    post_stats = Post.objects.filter(is_public=True).aggregate(
        total_posts=Count('id'),
        total_likes=Count('likes'),
        total_comments=Count('comments')
    )
    
    # Resource statistics
    resource_stats = Resource.objects.filter(is_public=True).aggregate(
        total_resources=Count('id'),
        total_downloads=Count('download_count')
    )
    
    # Tutorial statistics
    tutorial_stats = Tutorial.objects.filter(is_active=True).aggregate(
        total_tutorials=Count('id'),
        total_registrations=Count('registrations')
    )
    
    # Feedback statistics
    feedback_stats = Feedback.objects.aggregate(
        total_feedback=Count('id'),
        resolved_feedback=Count('id', filter=Q(status__in=['resolved', 'closed'])),
        pending_feedback=Count('id', filter=Q(status__in=['submitted', 'under_review']))
    )
    
    return Response({
        'success': True,
        'stats': {
            'users': user_stats,
            'posts': post_stats,
            'resources': resource_stats,
            'tutorials': tutorial_stats,
            'feedback': feedback_stats,
            'breakdown': {
                'by_department': list(users_by_department),
                'by_year': list(users_by_year),
                'by_zone': list(users_by_zone),
            }
        }
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_users_excel(request):
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MGSA Users"
    
    # Headers
    headers = [
        'ID', 'First Name', 'Middle Name', 'Last Name', 'Gender',
        'Zone', 'Woreda', 'College', 'Department', 'Year of Study',
        'Email', 'Student ID', 'Role', 'Executive Title', 'Date Joined'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    users = User.objects.filter(is_active=True).order_by('date_joined')
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.first_name)
        ws.cell(row=row, column=3, value=user.middle_name or '')
        ws.cell(row=row, column=4, value=user.last_name)
        ws.cell(row=row, column=5, value=user.gender)
        ws.cell(row=row, column=6, value=user.zone)
        ws.cell(row=row, column=7, value=user.woreda)
        ws.cell(row=row, column=8, value=user.college)
        ws.cell(row=row, column=9, value=user.department)
        ws.cell(row=row, column=10, value=user.year_of_study)
        ws.cell(row=row, column=11, value=user.email)
        ws.cell(row=row, column=12, value=user.student_id or '')
        ws.cell(row=row, column=13, value=user.role)
        ws.cell(row=row, column=14, value=user.executive_title or '')
        ws.cell(row=row, column=15, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mgsa_users.xlsx"'
    
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_users_pdf(request):
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Create PDF
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, height - 100, "MGSA Users Report")
    p.setFont("Helvetica", 12)
    p.drawString(100, height - 130, f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    
    # User statistics
    user_stats = User.objects.filter(is_active=True).aggregate(
        total=Count('id'),
        students=Count('id', filter=Q(role='Student')),
        executives=Count('id', filter=Q(role='Executive')),
        admins=Count('id', filter=Q(role='Admin'))
    )
    
    y_position = height - 180
    p.drawString(100, y_position, f"Total Users: {user_stats['total']}")
    y_position -= 20
    p.drawString(100, y_position, f"Students: {user_stats['students']}")
    y_position -= 20
    p.drawString(100, y_position, f"Executives: {user_stats['executives']}")
    y_position -= 20
    p.drawString(100, y_position, f"Admins: {user_stats['admins']}")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="mgsa_users_report.pdf"'
    
    return response

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def feedback_analytics(request):
    """Get feedback analytics for admin/executive"""
    if request.user.role not in ['Executive', 'Admin']:
        return Response({
            'success': False,
            'message': 'Access denied. Executive or Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Feedback by type
    feedback_by_type = Feedback.objects.values('feedback_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Feedback by status
    feedback_by_status = Feedback.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Recent feedback (last 7 days)
    recent_feedback = Feedback.objects.filter(
        created_at__gte=timezone.now() - timezone.timedelta(days=7)
    ).count()
    
    # Feedback resolution rate
    total_feedback = Feedback.objects.count()
    resolved_feedback = Feedback.objects.filter(
        status__in=['resolved', 'closed']
    ).count()
    
    resolution_rate = (resolved_feedback / total_feedback * 100) if total_feedback > 0 else 0
    
    return Response({
        'success': True,
        'analytics': {
            'total_feedback': total_feedback,
            'recent_feedback': recent_feedback,
            'resolution_rate': round(resolution_rate, 2),
            'by_type': list(feedback_by_type),
            'by_status': list(feedback_by_status),
        }
    })

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def submit_feedback(request):
    """Submit feedback from users"""
    from .serializers import FeedbackSerializer
    
    serializer = FeedbackSerializer(data=request.data)
    if serializer.is_valid():
        feedback = serializer.save(user=request.user)
        
        # Log user activity
        UserActivity.objects.create(
            user=request.user,
            activity_type='feedback_submit',
            description=f"Submitted feedback: {feedback.subject}",
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({
            'success': True,
            'message': 'Feedback submitted successfully!',
            'feedback': FeedbackSerializer(feedback).data
        }, status=status.HTTP_201_CREATED)
    
    return Response({
        'success': False,
        'message': 'Failed to submit feedback',
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def user_activity_logs(request):
    """Get user activity logs (admin only)"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Get activities from last 30 days
    activities = UserActivity.objects.filter(
        created_at__gte=timezone.now() - timezone.timedelta(days=30)
    ).select_related('user').order_by('-created_at')[:100]
    
    activity_data = []
    for activity in activities:
        activity_data.append({
            'user': activity.user.email,
            'activity_type': activity.get_activity_type_display(),
            'description': activity.description,
            'ip_address': activity.ip_address,
            'created_at': activity.created_at
        })
    
    return Response({
        'success': True,
        'activities': activity_data,
        'total_count': len(activity_data)
    })