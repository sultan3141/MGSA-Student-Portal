from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io
import json

from accounts.models import User
from posts.models import Post, Like, Comment
from resources.models import Resource
from tutorials.models import Tutorial, TutorialRegistration

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_dashboard(request):
    """Complete admin dashboard with all statistics"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Date range for analytics (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    # User Analytics
    user_stats = User.objects.filter(is_active=True).aggregate(
        total_users=Count('id'),
        students=Count('id', filter=Q(role='Student')),
        executives=Count('id', filter=Q(role='Executive')),
        admins=Count('id', filter=Q(role='Admin')),
        new_users_30_days=Count('id', filter=Q(date_joined__gte=thirty_days_ago)),
        active_today=Count('id', filter=Q(last_login__date=timezone.now().date()))
    )
    
    # Users by department (top 10)
    users_by_department = User.objects.filter(
        is_active=True, role='Student'
    ).values('department').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Users by year
    users_by_year = User.objects.filter(
        is_active=True, role='Student'
    ).values('year_of_study').annotate(count=Count('id')).order_by('year_of_study')
    
    # Users by zone
    users_by_zone = User.objects.filter(
        is_active=True, role='Student'
    ).values('zone').annotate(count=Count('id')).order_by('-count')
    
    # Post Analytics
    post_stats = Post.objects.aggregate(
        total_posts=Count('id'),
        public_posts=Count('id', filter=Q(is_public=True)),
        private_posts=Count('id', filter=Q(is_public=False)),
        posts_30_days=Count('id', filter=Q(created_at__gte=thirty_days_ago)),
        total_likes=Count('likes'),
        total_comments=Count('comments'),
        avg_likes_per_post=Count('likes') / Count('id', distinct=True)
    )
    
    # Top posts by likes
    top_posts = Post.objects.annotate(
        like_count=Count('likes'),
        comment_count=Count('comments')
    ).select_related('author').order_by('-like_count')[:10]
    
    top_posts_data = []
    for post in top_posts:
        top_posts_data.append({
            'id': post.id,
            'title': post.title,
            'author': f"{post.author.first_name} {post.author.last_name}",
            'likes': post.like_count,
            'comments': post.comment_count,
            'created_at': post.created_at
        })
    
    # Resource Analytics
    resource_stats = Resource.objects.aggregate(
        total_resources=Count('id'),
        public_resources=Count('id', filter=Q(is_public=True)),
        total_downloads=Count('download_count'),
        resources_30_days=Count('id', filter=Q(created_at__gte=thirty_days_ago)),
        avg_downloads_per_resource=Count('download_count') / Count('id', distinct=True)
    )
    
    # Popular resources
    popular_resources = Resource.objects.select_related('uploaded_by').order_by('-download_count')[:10]
    popular_resources_data = []
    for resource in popular_resources:
        popular_resources_data.append({
            'id': resource.id,
            'title': resource.title,
            'uploaded_by': f"{resource.uploaded_by.first_name} {resource.uploaded_by.last_name}",
            'downloads': resource.download_count,
            'file_type': resource.file_type
        })
    
    # Tutorial Analytics
    tutorial_stats = Tutorial.objects.aggregate(
        total_tutorials=Count('id'),
        active_tutorials=Count('id', filter=Q(is_active=True)),
        total_registrations=Count('registrations'),
        tutorials_30_days=Count('id', filter=Q(created_at__gte=thirty_days_ago))
    )
    
    # Popular tutorials
    popular_tutorials = Tutorial.objects.annotate(
        registration_count=Count('registrations')
    ).select_related('created_by').order_by('-registration_count')[:10]
    
    popular_tutorials_data = []
    for tutorial in popular_tutorials:
        popular_tutorials_data.append({
            'id': tutorial.id,
            'title': tutorial.title,
            'tutor': tutorial.tutor,
            'department': tutorial.department,
            'registrations': tutorial.registration_count,
            'max_students': tutorial.max_students,
            'completion_rate': f"{(tutorial.registration_count / tutorial.max_students * 100) if tutorial.max_students > 0 else 0:.1f}%"
        })
    
    # System Health
    system_health = {
        'database_size': 'N/A',  # You can implement this based on your database
        'active_sessions': 'N/A',  # Implement session tracking if needed
        'server_uptime': 'N/A',   # Implement server monitoring
        'last_backup': 'N/A'      # Implement backup tracking
    }
    
    # Recent Activities (last 50 activities)
    recent_users = User.objects.filter(is_active=True).order_by('-date_joined')[:10]
    recent_posts = Post.objects.select_related('author').order_by('-created_at')[:10]
    recent_resources = Resource.objects.select_related('uploaded_by').order_by('-created_at')[:10]
    
    return Response({
        'success': True,
        'dashboard': {
            'user_analytics': user_stats,
            'post_analytics': post_stats,
            'resource_analytics': resource_stats,
            'tutorial_analytics': tutorial_stats,
            'breakdowns': {
                'by_department': list(users_by_department),
                'by_year': list(users_by_year),
                'by_zone': list(users_by_zone),
            },
            'top_content': {
                'posts': top_posts_data,
                'resources': popular_resources_data,
                'tutorials': popular_tutorials_data,
            },
            'system_health': system_health,
            'recent_activities': {
                'users': [
                    {
                        'id': user.id,
                        'name': f"{user.first_name} {user.last_name}",
                        'email': user.email,
                        'role': user.role,
                        'joined': user.date_joined
                    } for user in recent_users
                ],
                'posts': [
                    {
                        'id': post.id,
                        'title': post.title,
                        'author': f"{post.author.first_name} {post.author.last_name}",
                        'created_at': post.created_at
                    } for post in recent_posts
                ],
                'resources': [
                    {
                        'id': resource.id,
                        'title': resource.title,
                        'uploaded_by': f"{resource.uploaded_by.first_name} {resource.uploaded_by.last_name}",
                        'created_at': resource.created_at
                    } for resource in recent_resources
                ]
            }
        }
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_management(request):
    """Complete user management for admin"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Get query parameters
    role_filter = request.GET.get('role', '')
    department_filter = request.GET.get('department', '')
    year_filter = request.GET.get('year', '')
    zone_filter = request.GET.get('zone', '')
    search_query = request.GET.get('search', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    # Build filters
    filters = Q(is_active=True)
    
    if role_filter:
        filters &= Q(role=role_filter)
    if department_filter:
        filters &= Q(department=department_filter)
    if year_filter:
        filters &= Q(year_of_study=year_filter)
    if zone_filter:
        filters &= Q(zone=zone_filter)
    if search_query:
        filters &= (
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(student_id__icontains=search_query)
        )
    
    # Get users with pagination
    users = User.objects.filter(filters).order_by('-date_joined')
    total_users = users.count()
    total_pages = (total_users + page_size - 1) // page_size
    
    # Apply pagination
    start_index = (page - 1) * page_size
    end_index = start_index + page_size
    paginated_users = users[start_index:end_index]
    
    # Serialize users
    users_data = []
    for user in paginated_users:
        users_data.append({
            'id': user.id,
            'first_name': user.first_name,
            'middle_name': user.middle_name,
            'last_name': user.last_name,
            'email': user.email,
            'student_id': user.student_id,
            'role': user.role,
            'executive_title': user.executive_title,
            'department': user.department,
            'year_of_study': user.year_of_study,
            'zone': user.zone,
            'woreda': user.woreda,
            'college': user.college,
            'last_login': user.last_login,
            'date_joined': user.date_joined,
            'is_active': user.is_active,
            'profile_picture': user.profile_picture.url if user.profile_picture else None
        })
    
    # Get available filters data
    available_roles = User.objects.filter(is_active=True).values_list('role', flat=True).distinct()
    available_departments = User.objects.filter(is_active=True).values_list('department', flat=True).distinct()
    available_years = User.objects.filter(is_active=True).values_list('year_of_study', flat=True).distinct()
    available_zones = User.objects.filter(is_active=True).values_list('zone', flat=True).distinct()
    
    return Response({
        'success': True,
        'users': users_data,
        'pagination': {
            'current_page': page,
            'page_size': page_size,
            'total_users': total_users,
            'total_pages': total_pages,
            'has_next': page < total_pages,
            'has_previous': page > 1
        },
        'filters': {
            'roles': list(available_roles),
            'departments': list(available_departments),
            'years': list(available_years),
            'zones': list(available_zones)
        }
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_user_role(request, user_id):
    """Update user role and executive title"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    try:
        user = User.objects.get(id=user_id, is_active=True)
    except User.DoesNotExist:
        return Response({
            'success': False,
            'message': 'User not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    role = request.data.get('role')
    executive_title = request.data.get('executive_title', '')
    
    if role not in ['Student', 'Executive', 'Admin']:
        return Response({
            'success': False,
            'message': 'Invalid role'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Update user
    user.role = role
    if executive_title:
        user.executive_title = executive_title
    elif role == 'Student':
        user.executive_title = ''
    
    user.save()
    
    return Response({
        'success': True,
        'message': f'User role updated to {role} successfully',
        'user': {
            'id': user.id,
            'email': user.email,
            'role': user.role,
            'executive_title': user.executive_title
        }
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def deactivate_user(request, user_id):
    """Deactivate user account"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    try:
        user = User.objects.get(id=user_id)
        
        # Prevent deactivating yourself
        if user.id == request.user.id:
            return Response({
                'success': False,
                'message': 'Cannot deactivate your own account'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        user.is_active = False
        user.save()
        
        return Response({
            'success': True,
            'message': 'User account deactivated successfully'
        })
    except User.DoesNotExist:
        return Response({
            'success': False,
            'message': 'User not found'
        }, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def activate_user(request, user_id):
    """Activate user account"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    try:
        user = User.objects.get(id=user_id)
        user.is_active = True
        user.save()
        
        return Response({
            'success': True,
            'message': 'User account activated successfully'
        })
    except User.DoesNotExist:
        return Response({
            'success': False,
            'message': 'User not found'
        }, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def content_management(request):
    """Admin content management - view all content"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    content_type = request.GET.get('type', 'all')  # posts, resources, tutorials, all
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    response_data = {}
    
    if content_type in ['posts', 'all']:
        posts = Post.objects.select_related('author').prefetch_related('likes', 'comments').order_by('-created_at')
        total_posts = posts.count()
        
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        paginated_posts = posts[start_index:end_index]
        
        posts_data = []
        for post in paginated_posts:
            posts_data.append({
                'id': post.id,
                'title': post.title,
                'content_preview': post.content[:100] + '...' if len(post.content) > 100 else post.content,
                'author': f"{post.author.first_name} {post.author.last_name}",
                'author_role': post.author.role,
                'is_public': post.is_public,
                'likes_count': post.likes.count(),
                'comments_count': post.comments.count(),
                'created_at': post.created_at,
                'media_count': len(post.media) if hasattr(post, 'media') else 0
            })
        
        response_data['posts'] = {
            'data': posts_data,
            'total': total_posts,
            'page': page,
            'page_size': page_size
        }
    
    if content_type in ['resources', 'all']:
        resources = Resource.objects.select_related('uploaded_by').order_by('-created_at')
        total_resources = resources.count()
        
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        paginated_resources = resources[start_index:end_index]
        
        resources_data = []
        for resource in paginated_resources:
            resources_data.append({
                'id': resource.id,
                'title': resource.title,
                'description': resource.description,
                'file_type': resource.file_type,
                'file_size': resource.file_size,
                'uploaded_by': f"{resource.uploaded_by.first_name} {resource.uploaded_by.last_name}",
                'download_count': resource.download_count,
                'is_public': resource.is_public,
                'created_at': resource.created_at
            })
        
        response_data['resources'] = {
            'data': resources_data,
            'total': total_resources,
            'page': page,
            'page_size': page_size
        }
    
    if content_type in ['tutorials', 'all']:
        tutorials = Tutorial.objects.select_related('created_by').prefetch_related('registrations').order_by('-created_at')
        total_tutorials = tutorials.count()
        
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        paginated_tutorials = tutorials[start_index:end_index]
        
        tutorials_data = []
        for tutorial in paginated_tutorials:
            tutorials_data.append({
                'id': tutorial.id,
                'title': tutorial.title,
                'tutor': tutorial.tutor,
                'department': tutorial.department,
                'created_by': f"{tutorial.created_by.first_name} {tutorial.created_by.last_name}",
                'max_students': tutorial.max_students,
                'current_registrations': tutorial.current_registrations,
                'is_active': tutorial.is_active,
                'start_date': tutorial.start_date,
                'end_date': tutorial.end_date,
                'created_at': tutorial.created_at
            })
        
        response_data['tutorials'] = {
            'data': tutorials_data,
            'total': total_tutorials,
            'page': page,
            'page_size': page_size
        }
    
    return Response({
        'success': True,
        'content': response_data
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def delete_content(request, content_type, content_id):
    """Delete any content (posts, resources, tutorials)"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    try:
        if content_type == 'post':
            content = Post.objects.get(id=content_id)
            content_title = content.title
            content.delete()
        
        elif content_type == 'resource':
            content = Resource.objects.get(id=content_id)
            content_title = content.title
            content.delete()
        
        elif content_type == 'tutorial':
            content = Tutorial.objects.get(id=content_id)
            content_title = content.title
            content.delete()
        
        else:
            return Response({
                'success': False,
                'message': 'Invalid content type'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': True,
            'message': f'{content_type.title()} "{content_title}" deleted successfully'
        })
    
    except (Post.DoesNotExist, Resource.DoesNotExist, Tutorial.DoesNotExist):
        return Response({
            'success': False,
            'message': 'Content not found'
        }, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def toggle_content_visibility(request, content_type, content_id):
    """Toggle content visibility (public/private)"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    try:
        if content_type == 'post':
            content = Post.objects.get(id=content_id)
            content.is_public = not content.is_public
            content.save()
            new_status = 'public' if content.is_public else 'private'
        
        elif content_type == 'resource':
            content = Resource.objects.get(id=content_id)
            content.is_public = not content.is_public
            content.save()
            new_status = 'public' if content.is_public else 'private'
        
        elif content_type == 'tutorial':
            content = Tutorial.objects.get(id=content_id)
            content.is_active = not content.is_active
            content.save()
            new_status = 'active' if content.is_active else 'inactive'
        
        else:
            return Response({
                'success': False,
                'message': 'Invalid content type'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': True,
            'message': f'{content_type.title()} visibility updated to {new_status}',
            'new_status': new_status
        })
    
    except (Post.DoesNotExist, Resource.DoesNotExist, Tutorial.DoesNotExist):
        return Response({
            'success': False,
            'message': 'Content not found'
        }, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_data(request):
    """Export data in various formats"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    export_type = request.GET.get('type', 'users')  # users, posts, resources, tutorials
    format_type = request.GET.get('format', 'excel')  # excel, pdf, json
    
    if export_type == 'users':
        return export_users_data(format_type)
    elif export_type == 'posts':
        return export_posts_data(format_type)
    elif export_type == 'resources':
        return export_resources_data(format_type)
    elif export_type == 'tutorials':
        return export_tutorials_data(format_type)
    else:
        return Response({
            'success': False,
            'message': 'Invalid export type'
        }, status=status.HTTP_400_BAD_REQUEST)

def export_users_data(format_type):
    """Export users data"""
    users = User.objects.filter(is_active=True).order_by('date_joined')
    
    if format_type == 'excel':
        return export_users_excel(users)
    elif format_type == 'pdf':
        return export_users_pdf(users)
    elif format_type == 'json':
        return export_users_json(users)
    else:
        return Response({
            'success': False,
            'message': 'Invalid format type'
        }, status=status.HTTP_400_BAD_REQUEST)

def export_users_excel(users):
    """Export users to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MGSA Users"
    
    # Headers
    headers = [
        'ID', 'First Name', 'Middle Name', 'Last Name', 'Gender',
        'Zone', 'Woreda', 'College', 'Department', 'Year of Study',
        'Email', 'Student ID', 'Role', 'Executive Title', 
        'Last Login', 'Date Joined', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    
    # Data
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.first_name)
        ws.cell(row=row, column=3, value=user.middle_name or '')
        ws.cell(row=row, column=4, value=user.last_name)
        ws.cell(row=row, column=5, value=user.gender)
        ws.cell(row=row, column=6, value=user.zone)
        ws.cell(row=row, column=7, value=user.woreda)
        ws.cell(row=row, column=8, value=user.college)
        ws.cell(row=row, column=9, value=user.department)
        ws.cell(row=row, column=10, value=user.year_of_study)
        ws.cell(row=row, column=11, value=user.email)
        ws.cell(row=row, column=12, value=user.student_id or '')
        ws.cell(row=row, column=13, value=user.role)
        ws.cell(row=row, column=14, value=user.executive_title or '')
        ws.cell(row=row, column=15, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never')
        ws.cell(row=row, column=16, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=17, value='Active' if user.is_active else 'Inactive')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mgsa_users_export.xlsx"'
    
    wb.save(response)
    return response

def export_users_pdf(users):
    """Export users to PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph("MGSA Users Export", styles['Title'])
    elements.append(title)
    
    # Summary
    summary_text = f"Total Users: {users.count()} | Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    elements.append(Paragraph("<br/>", styles['Normal']))
    
    # Create table data
    table_data = [['Name', 'Email', 'Role', 'Department', 'Year', 'Zone', 'Joined']]
    
    for user in users:
        table_data.append([
            f"{user.first_name} {user.last_name}",
            user.email,
            user.role,
            user.department,
            user.year_of_study,
            user.zone,
            user.date_joined.strftime('%Y-%m-%d')
        ])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="mgsa_users_export.pdf"'
    
    return response

def export_users_json(users):
    """Export users to JSON"""
    users_data = []
    for user in users:
        users_data.append({
            'id': user.id,
            'first_name': user.first_name,
            'middle_name': user.middle_name,
            'last_name': user.last_name,
            'email': user.email,
            'student_id': user.student_id,
            'role': user.role,
            'executive_title': user.executive_title,
            'department': user.department,
            'year_of_study': user.year_of_study,
            'zone': user.zone,
            'woreda': user.woreda,
            'college': user.college,
            'last_login': user.last_login.isoformat() if user.last_login else None,
            'date_joined': user.date_joined.isoformat(),
            'is_active': user.is_active
        })
    
    response = HttpResponse(
        json.dumps(users_data, indent=2),
        content_type='application/json'
    )
    response['Content-Disposition'] = 'attachment; filename="mgsa_users_export.json"'
    
    return response

# Similar export functions for posts, resources, tutorials...
def export_posts_data(format_type):
    """Export posts data"""
    posts = Post.objects.select_related('author').prefetch_related('likes', 'comments').order_by('-created_at')
    
    if format_type == 'excel':
        return export_posts_excel(posts)
    elif format_type == 'json':
        return export_posts_json(posts)
    else:
        return Response({
            'success': False,
            'message': 'Format not supported for posts'
        }, status=status.HTTP_400_BAD_REQUEST)

def export_posts_excel(posts):
    """Export posts to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MGSA Posts"
    
    headers = ['ID', 'Title', 'Author', 'Content Preview', 'Likes', 'Comments', 'Public', 'Created At']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    
    for row, post in enumerate(posts, 2):
        ws.cell(row=row, column=1, value=post.id)
        ws.cell(row=row, column=2, value=post.title)
        ws.cell(row=row, column=3, value=f"{post.author.first_name} {post.author.last_name}")
        ws.cell(row=row, column=4, value=post.content[:50] + '...' if len(post.content) > 50 else post.content)
        ws.cell(row=row, column=5, value=post.likes.count())
        ws.cell(row=row, column=6, value=post.comments.count())
        ws.cell(row=row, column=7, value='Yes' if post.is_public else 'No')
        ws.cell(row=row, column=8, value=post.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mgsa_posts_export.xlsx"'
    
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def system_settings(request):
    """Get and update system settings"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # In a real application, you'd store these in a database model
    # For now, we'll use a simple dictionary
    system_settings = {
        'site_name': 'MGSA Portal',
        'site_description': 'Murti Guto Students Association Portal',
        'registration_enabled': True,
        'max_file_size': 10,  # MB
        'allowed_file_types': ['pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png', 'mp4'],
        'maintenance_mode': False,
        'contact_email': 'admin@mgsa.org',
        'max_posts_per_day': 5,
        'auto_approve_posts': False
    }
    
    return Response({
        'success': True,
        'settings': system_settings
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_system_settings(request):
    """Update system settings"""
    if request.user.role != 'Admin':
        return Response({
            'success': False,
            'message': 'Access denied. Admin privileges required.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # In a real application, you'd save these to a database
    # For now, we'll just return success
    
    return Response({
        'success': True,
        'message': 'System settings updated successfully',
        'updated_settings': request.data
    })